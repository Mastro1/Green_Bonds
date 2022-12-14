import requests
import openpyxl

wb = openpyxl.load_workbook('ICMA-Sustainable-Bonds-Database-151022.xlsx')


def download_pdf(sheet):
    ws = wb[sheet]
    last_raw = ws.max_row
    print(ws.cell(row=3, column=6).hyperlink.target)
    list_link = []
    for bond in range(3, last_raw):
        try:
            link = ws.cell(row=bond, column=6).hyperlink.target
            name = ws.cell(row=bond, column=1).value
            list_link.append((name, link))
        except:
            pass
    i = 0
    errors = []
    for link in list_link:
        i += 1
        try:

            print("Downloading file: ", i)
            response = requests.get(link[1])
            file = open("files/{}/" + "{}.pdf".format(sheet, link[0]), "wb")
            file.write(response.content)
            file.close()
            print("File ", i, " downloaded")
        except:
            print("file {} had a problem".format(i))
            print(link)
            errors.append(link)

    print("All PDF files downloaded")
    print("Except following errors:")
    print(errors)
    print("total errors: {}".format(len(errors)))


for sheet in wb.sheetnames:
    download_pdf(sheet)